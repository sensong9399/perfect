# -*- coding: utf-8 -*-
import time
import json
import MySQLdb
import openpyxl as xl
from tqdm import tqdm
import sys

def readconf(configfile="/home/client/boxconfig/config.json"):
    #读取系统配置文件
    f = open(configfile)
    content = f.read()
    conf = json.loads(content)
    f.close()
    return conf


def mysqlexe(mysql_host, database, username, password, sqlstr):
    #向mysql插值
    try:
        db = MySQLdb.connect(mysql_host,username, password,database)
        cursor = db.cursor()
        try:
            cursor.execute(sqlstr)
            data = cursor.fetchall()
            return data
            #db.commit()
            cursor.close()
            db.close()

        except Exception, e:
            printlog("update tag's range, mysql execute error! " + str(e))
            print(e)
    except:
        printlog("mysql setting error! " + str(e))
        print(e)




def printlog(logtext,logtime = ""):
    #记录并打印日志文本，若未输入时间使用当前时间
    if logtime == "":
        logtime = time.localtime()
    timestr = str(time.strftime("%Y-%m-%d %H:%M:%S", logtime))
    logfile = open("logfile","a+")
    readinglog = timestr + ", " + logtext + "\n"
    logfile.write(readinglog)
    logfile.close


if __name__ == "__main__":
    if len(sys.argv)<3:
        print("Please specify the start and end date")
        exit()
    elif sys.argv[2]<sys.argv[1]:
        print("End date < start date ?!")
        exit()
    start_date = sys.argv[1]
    try:
        st=time.strptime(start_date+" 00:00:00","%Y-%m-%d %H:%M:%S")
        start_time=int(time.mktime(st))
    except Exception,e:
        print("start date format error!")
        exit()
    end_date = sys.argv[2]
    try:
        ed=time.strptime(end_date+" 00:00:00","%Y-%m-%d %H:%M:%S")
        end_time=int(time.mktime(ed))
    except Exception,e:
        print("end date format error!")
        exit()
    sqlstr = "select item_id,max(value+0) max_v, min(value+0) min_v,max(value+0)-min(value+0) scope \
            from item_data \
            where long_time>="+str(start_time)+" and long_time<="+str(end_time)+\
            " group by item_id;"
    configj = readconf()
    data = mysqlexe(configj["mysql_host"],configj["database"],configj["username"],configj["password"], sqlstr)
    wb = xl.load_workbook("/home/client/boxconfig/tags.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    for i in tqdm(range(1,ws.max_row+1)):
        for r in data:
            if ws["D"+str(i)].value == r[0]:
                if r[0]=='[GOLDEN.LPD_OPC]LT000A_Sim.Val':
                    print(ws["D"+str(i)],r[0],r[1],r[2],r[3])
                ws["N"+str(i)] = r[1]
                ws["O"+str(i)] = r[2]
                ws["P"+str(i)] = r[3]
                #print(ws["D"+str(i)].value,r[0],r[1],r[2],r[3])
    wb.save("/home/client/boxconfig/tags.xlsx")
    for i in tqdm(range(1,ws.max_row)):
        if ws["P"+str(i)].value == None:
            ws["N"+str(i)] = 0.0
            ws["O"+str(i)] = 0.0
            ws["P"+str(i)] = 0.0001
        elif ws["P"+str(i)].value==0:
            ws["P"+str(i)] = 0.0001

    wb.save("/home/client/boxconfig/tags.xlsx")
